"""Export all evaluation ground truth (GT) datasets to a single Excel file.

Each query file under eval/data/queries/ becomes a sheet in the workbook,
with one row per query including:

- query text, type, level, category
- relevant_docs (GT IDs)
- relevant_answer (resolved to human-readable title + content snippet)

The GT ID → answer resolution uses the corresponding graph sqlite file,
so you can see the actual answer text alongside the cryptic doc ID.

Usage::

    uv run python eval/scripts/export_gt_to_excel.py
    → writes eval/data/gt_datasets.xlsx
"""

from __future__ import annotations

import json
import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

REPO_ROOT = Path(__file__).resolve().parents[2]
QUERIES_DIR = REPO_ROOT / "eval" / "data" / "queries"
OUTPUT_PATH = REPO_ROOT / "eval" / "data" / "gt_datasets.xlsx"

# Map query-file name → graph sqlite file for resolving GT IDs to text.
GRAPH_MAP = {
    "krra": "krra_graph.sqlite",
    "krra_hard": "krra_graph.sqlite",
    "krra_graph": "krra_graph.sqlite",
    "krra_multihop": "krra_graph.sqlite",
    "assort": "assort_graph.sqlite",
    "assort_hard": "assort_graph.sqlite",
    "x2bee": "x2bee_graph.sqlite",
    "x2bee_hard": "x2bee_graph.sqlite",
}


def _load_resolver(graph_path: Path) -> dict[str, str]:
    """Build a map from GT-style ID to human-readable answer.

    Keys cover both:
    - node title (for structured data like "products:12800000")
    - properties.doc_id (for document data with hash IDs)
    """
    if not graph_path.exists():
        return {}

    resolver: dict[str, str] = {}
    try:
        conn = sqlite3.connect(str(graph_path))
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT title, content, properties_json FROM syn_nodes"
        ).fetchall()
        for r in rows:
            title = r["title"] or ""
            raw_content = (r["content"] or "").replace("\n", " ").strip()

            # Strip redundant "table_name: " prefix from structured content
            # so the answer column is readable.
            if ":" in title:
                tbl = title.split(":", 1)[0]
                prefix = f"{tbl}: "
                if raw_content.startswith(prefix):
                    raw_content = raw_content[len(prefix):]

            # Cap content length for readability in Excel cells
            content_snippet = raw_content[:150]
            if len(raw_content) > 150:
                content_snippet += "…"

            # For structured nodes: title already shows "table:pk", so just
            # show the content (which starts with the name). For document
            # nodes: title is the doc/chunk title; prepend it.
            try:
                props = json.loads(r["properties_json"] or "{}")
            except json.JSONDecodeError:
                props = {}

            is_structured = bool(props.get("_table_name"))
            if is_structured:
                summary = content_snippet or title
            else:
                # Document node — show title + content preview
                summary = f"{title}  ▸  {content_snippet}" if content_snippet else title

            # Key by title (structured: "products:12800000")
            if title:
                resolver[title] = summary

            # Key by properties.doc_id (documents: "0346542e...")
            did = props.get("doc_id", "")
            if did:
                resolver[str(did)] = summary

        conn.close()
    except Exception as exc:
        print(f"  ⚠ resolver failed for {graph_path.name}: {exc}")

    return resolver


HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)


def _flatten_query(q: dict, resolver: dict[str, str] | None = None) -> dict:
    """Normalize a query dict to a flat set of columns.

    When ``resolver`` is provided, the `relevant_answer` column is
    populated with human-readable title+content for each GT ID.
    """
    relevant = q.get("relevant_docs") or q.get("answer_ids") or []
    if isinstance(relevant, dict):
        relevant = list(relevant.keys())

    # Resolve GT IDs to readable answers (title + content snippet)
    resolved_lines: list[str] = []
    if resolver is not None:
        for i, rid in enumerate(relevant[:20], start=1):  # cap at 20 for readability
            key = str(rid)
            answer = resolver.get(key, "")
            if not answer:
                # Try stripping chunk suffix ("#1", "#2") for doc lookup
                base = key.rsplit(" #", 1)[0]
                answer = resolver.get(base, "")
            if answer:
                resolved_lines.append(f"{i}. {answer}")
            else:
                resolved_lines.append(f"{i}. [{key}] (not found)")
        if len(relevant) > 20:
            resolved_lines.append(f"... +{len(relevant) - 20} more")

    return {
        "qid": q.get("qid", q.get("query_id", "")),
        "query": q.get("query", q.get("question", "")),
        "type": q.get("type", q.get("query_type", "")),
        "level": q.get("level", q.get("difficulty", "")),
        "category": q.get("category", ""),
        "description": q.get("description", ""),
        "relevant_count": len(relevant),
        "relevant_answer": "\n".join(resolved_lines),
        "relevant_docs": "\n".join(str(x) for x in relevant),
    }


def _write_sheet(
    wb: Workbook,
    name: str,
    meta: dict,
    queries: list[dict],
    resolver: dict[str, str] | None = None,
) -> None:
    ws = wb.create_sheet(title=name[:31])  # Excel sheet name limit

    # Metadata header (first rows)
    ws["A1"] = "Dataset"
    ws["B1"] = name
    ws["A2"] = "Description"
    ws["B2"] = meta.get("description", "")
    ws["A3"] = "id_field"
    ws["B3"] = meta.get("id_field", "doc_id")
    ws["A4"] = "Total queries"
    ws["B4"] = len(queries)
    ws["A5"] = "Answer resolved?"
    ws["B5"] = "YES — see relevant_answer column" if resolver else "NO graph found"

    for row in range(1, 6):
        ws[f"A{row}"].font = Font(bold=True)

    # Column headers
    columns = ["qid", "query", "type", "level", "category", "description",
               "relevant_count", "relevant_answer", "relevant_docs"]
    header_row = 7
    for i, col in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=i, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for r, q in enumerate(queries, start=header_row + 1):
        flat = _flatten_query(q, resolver=resolver)
        for c, col in enumerate(columns, start=1):
            cell = ws.cell(row=r, column=c, value=flat.get(col, ""))
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Column widths — wider for answer column
    widths = {
        "A": 8, "B": 45, "C": 18, "D": 8, "E": 22, "F": 42,
        "G": 14, "H": 70, "I": 45,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Row heights (to accommodate wrap)
    for r in range(header_row + 1, header_row + 1 + len(queries)):
        ws.row_dimensions[r].height = 80

    # Freeze header
    ws.freeze_panes = f"A{header_row + 1}"


def _write_summary(wb: Workbook, stats: list[dict]) -> None:
    ws = wb.create_sheet(title="Summary", index=0)

    ws["A1"] = "Synaptic Memory — Evaluation GT Datasets"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")

    ws["A3"] = "Generated from"
    ws["B3"] = "eval/data/queries/*.json"
    ws["A3"].font = Font(bold=True)

    header_row = 5
    headers = ["Dataset", "Description", "Queries", "id_field", "Language"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=i, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for r, s in enumerate(stats, start=header_row + 1):
        for c, key in enumerate(["dataset", "description", "queries", "id_field", "language"], start=1):
            cell = ws.cell(row=r, column=c, value=s.get(key, ""))
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {"A": 22, "B": 60, "C": 10, "D": 14, "E": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = f"A{header_row + 1}"


def _guess_language(name: str, queries: list[dict]) -> str:
    # Sample a few queries
    text = " ".join(str(q.get("query", "")) for q in queries[:5])
    if any(ord(c) > 0x3000 and ord(c) < 0xD7AF for c in text):
        return "KO"
    return "EN"


def main() -> None:
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    if default is not None:
        wb.remove(default)

    # Pre-load resolvers for each unique graph (avoid re-loading per sheet)
    resolver_cache: dict[str, dict[str, str]] = {}

    files = sorted(QUERIES_DIR.glob("*.json"))
    stats: list[dict] = []

    for path in files:
        with open(path, encoding="utf-8") as f:
            try:
                data = json.load(f)
            except json.JSONDecodeError as exc:
                print(f"  ⚠ skip {path.name}: {exc}")
                continue

        queries = data.get("queries", [])
        if not isinstance(queries, list) or not queries:
            print(f"  ⚠ skip {path.name}: no queries list")
            continue

        name = path.stem
        meta = {
            "description": data.get("description", ""),
            "id_field": data.get("id_field", "doc_id"),
        }

        # Load resolver for this dataset's graph
        resolver: dict[str, str] | None = None
        graph_file = GRAPH_MAP.get(name)
        if graph_file:
            if graph_file not in resolver_cache:
                graph_path = REPO_ROOT / "eval" / "data" / graph_file
                resolver_cache[graph_file] = _load_resolver(graph_path)
                print(f"  📖 loaded {graph_file}: {len(resolver_cache[graph_file])} entries")
            resolver = resolver_cache[graph_file]

        _write_sheet(wb, name, meta, queries, resolver=resolver)
        stats.append({
            "dataset": name,
            "description": meta["description"][:100],
            "queries": len(queries),
            "id_field": meta["id_field"],
            "language": _guess_language(name, queries),
        })
        resolved = " (with answers)" if resolver else ""
        print(f"  ✓ {name}: {len(queries)} queries{resolved}")

    _write_summary(wb, stats)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"\nSaved → {OUTPUT_PATH}")
    print(f"Total datasets: {len(stats)}")


if __name__ == "__main__":
    main()
